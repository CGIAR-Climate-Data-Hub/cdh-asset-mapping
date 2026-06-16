"""
ingest.py
---------
Reads all centre Excel submission files, normalises key fields, applies the
merge log, and writes data/assets.json.

Usage:
    python src/ingest.py

Output:
    data/normalized/assets.json
"""

import json
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).parent.parent
SUBMISSIONS_DIR = ROOT / "data" / "submissions"
MERGE_LOG_PATH  = ROOT / "data" / "merge_log.json"
OUTPUT_PATH     = ROOT / "data" / "normalized" / "assets.json"

# ---------------------------------------------------------------------------
# Centre configuration
# ---------------------------------------------------------------------------
CENTRE_FILES = {
    "AfricaRice":           "AfricaRice.xlsx",
    "Alliance":             "Alliance.xlsx",
    "CIFOR_ICRAF":          "CIFOR_ICRAF.xlsx",
    "CIP":                  "CIP.xlsx",
    "ICRISAT":              "ICRISAT.xlsx",
    "IFPRI":                "IFPRI.xlsx",
    "IITA":                 "IITA.xlsx",
    "ILRI":                 "ILRI.xlsx",
    "IRRI":                 "IRRI.xlsx",
    "IWMI":                 "IWMI.xlsx",
    "WorldFish":            "WorldFish.xlsx",
}

HUB_FUNDED = {"Alliance", "IITA", "ILRI", "IFPRI", "IWMI", "WorldFish"}

# Expected sheet names (some centres may use slight variations)
SHEET_A = "A Identity"
SHEET_B = "B Structure"
SHEET_C = "C SpatioTemp"
SHEET_D = "D Thematic"
SHEET_E = "E Context & Use"
SHEET_FG = "F & G Assess & Nominate"

# ---------------------------------------------------------------------------
# Normalisation maps
# ---------------------------------------------------------------------------
DOMAIN_NORM = {
    # Hazard
    "hazard":                               "Hazard",
    "climate hazard":                       "Hazard",
    "climatology":                          "Hazard",
    "rainfall seasonality":                 "Hazard",
    # Hazard / Climate Services
    "climate services":                     "Hazard / Climate Services",
    "climate service":                      "Hazard / Climate Services",
    "agrometeorology":                      "Hazard / Climate Services",
    "agro-meteorology":                     "Hazard / Climate Services",
    # Exposure
    "exposure":                             "Exposure",
    "agricultural systems":                 "Exposure",
    "land use":                             "Exposure",
    # Sensitivity
    "sensitivity":                          "Sensitivity",
    "food safety risk":                     "Sensitivity",
    "disease risk":                         "Sensitivity",
    # Adaptive Capacity
    "adaptive capacity":                    "Adaptive Capacity",
    # Adaptation Analytics
    "adaptation analytics":                 "Adaptation Analytics",
    "adaptation analytic":                  "Adaptation Analytics",
    "impact":                               "Adaptation Analytics",
    "agronomy and climate":                 "Adaptation Analytics",
    "agronomy":                             "Adaptation Analytics",
    # Mitigation
    "mitigation":                           "Mitigation",
    "ghg":                                  "Mitigation",
    "greenhouse gas":                       "Mitigation",
    # Climate Policy / Finance
    "policy":                               "Climate Policy / Finance",
    "finance":                              "Climate Policy / Finance",
    # Multi-domain
    "multi":                                "Multi-domain",
    "one health":                           "Multi-domain",
    "food security":                        "Multi-domain",
    "food systems":                         "Multi-domain",
    "biodiversity":                         "Multi-domain",
    "pest and disease":                     "Multi-domain",
    "breeding":                             "Multi-domain",
}

HYBRID_PATTERNS = [
    (r"adaptation.*(mitigation|ghg)",       "Adaptation Analytics / Mitigation"),
    (r"(mitigation|ghg).*adaptation",       "Adaptation Analytics / Mitigation"),
    (r"sensitivity.*adaptation",            "Sensitivity / Adaptation Analytics"),
    (r"adaptation.*sensitivity",            "Sensitivity / Adaptation Analytics"),
    (r"adaptive.*adaptation",               "Adaptive Capacity / Adaptation Analytics"),
    (r"adaptation.*adaptive",               "Adaptive Capacity / Adaptation Analytics"),
    (r"hazard.*exposure",                   "Hazard / Exposure"),
    (r"exposure.*hazard",                   "Hazard / Exposure"),
    (r"hazard.*climate.service",            "Hazard / Climate Services"),
]

TYPE_NORM = {
    "cgiar-produced":   "CGIAR-produced",
    "cgiar produced":   "CGIAR-produced",
    "cgiar":            "CGIAR-produced",
    "co-produced":      "Co-produced",
    "co produced":      "Co-produced",
    "external":         "External",
    "open":             "CGIAR-produced",  # Alliance used "Open" for their own assets
}

GEO_NORM_MAP = {
    # Africa
    "africa":               "Africa",
    "sub-saharan":          "Africa",
    "west africa":          "Africa",
    "east africa":          "Africa",
    "southern africa":      "Africa",
    "nigeria":              "Africa",
    "ghana":                "Africa",
    "kenya":                "Africa",
    "ethiopia":             "Africa",
    "zambia":               "Africa",
    "tanzania":             "Africa",
    "cameroon":             "Africa",
    "rwanda":               "Africa",
    "gambia":               "Africa",
    "mali":                 "Africa",
    "benin":                "Africa",
    "senegal":              "Africa",
    "inland valley":        "Africa",     # AfricaRice agro-ecosystem descriptor
    "rice system":          "Africa",     # AfricaRice
    "farming communit":     "Africa",     # AfricaRice
    "irrigated rice":       "Africa",     # AfricaRice
    "diverse rice":         "Africa",     # AfricaRice
    # Global
    "global":               "Global",
    "worldwide":            "Global",
    "world":                "Global",
    "multi-country":        "Global",
    "multiple":             "Global",
    "varies":               "Global",
    "vary by":              "Global",
    "flexible":             "Global",
    "prototype":            "Global",
    "continental":          "Global",
    "specific region":      "Global",    # ICRISAT "Specific regions"
    # Asia / South & SE Asia
    "asia":                 "Asia / South & SE Asia",
    "south asia":           "Asia / South & SE Asia",
    "southeast asia":       "Asia / South & SE Asia",
    "philippines":          "Asia / South & SE Asia",
    "phl":                  "Asia / South & SE Asia",
    "tha":                  "Asia / South & SE Asia",
    "vnm":                  "Asia / South & SE Asia",
    "khm":                  "Asia / South & SE Asia",
    "ind":                  "Asia / South & SE Asia",
    "idn":                  "Asia / South & SE Asia",
    "bgd":                  "Asia / South & SE Asia",
    "india":                "Asia / South & SE Asia",
    "bangladesh":           "Asia / South & SE Asia",
    "indonesia":            "Asia / South & SE Asia",
    "rice-growing":         "Asia / South & SE Asia",
    "irri hq":              "Asia / South & SE Asia",   # IRRI HQ = Los Baños, Philippines
    # Latin America & Caribbean
    "latin america":        "Latin America & Caribbean",
    "lac":                  "Latin America & Caribbean",
    "colombia":             "Latin America & Caribbean",
    "columbia":             "Latin America & Caribbean",  # common typo
    "peru":                 "Latin America & Caribbean",
    "central america":      "Latin America & Caribbean",
    "honduras":             "Latin America & Caribbean",
    # Multi-regional handled by regex below
}


def normalise_domain(raw: str) -> str:
    if not raw:
        return "Not specified"
    low = raw.lower().strip()
    # Check hybrid patterns first
    for pattern, label in HYBRID_PATTERNS:
        if re.search(pattern, low):
            return label
    # Check keyword map
    for kw, label in DOMAIN_NORM.items():
        if kw in low:
            return label
    return "Multi-domain"


def normalise_type(raw: str) -> str:
    if not raw:
        return "Not specified"
    low = str(raw).lower().strip()
    # Co-produced check first: "CGIAR-produced / Co-produced" and
    # "CGIAR-produced+ External" should not collapse to CGIAR-produced.
    if "co-produced" in low or "co produced" in low or "coproduced" in low:
        return "Co-produced"
    for kw, label in TYPE_NORM.items():
        if kw in low:
            return label
    # Unmappable non-empty value = a misplaced data-product description, not an
    # ownership class. Do not leak it into the category axis.
    return "Uncategorised"


def normalise_geo(raw: str) -> str:
    if not raw:
        return "Not specified"
    low = str(raw).lower()
    # Multi-regional patterns first (regex)
    for pattern, label in [
        (r"africa.*(asia|lac|latin)", "Multi-regional"),
        (r"(asia|lac|latin).*africa",  "Multi-regional"),
        (r"asia.*(lac|latin)",         "Multi-regional"),
    ]:
        if re.search(pattern, low):
            return label
    for kw, label in GEO_NORM_MAP.items():
        if kw in low:
            return label
    return "Not specified"


# ---------------------------------------------------------------------------
# Ordinal quality scores (Decision Relevance, Reuse Potential, Technical
# Readiness, Contemporary Validity, Sustainability)
#
# Submissions use a Low..Very High scale with inconsistent spelling
# ("Medium–High", "Medium-High", "medium-high", "high"). normalise_ordinal
# collapses these to a clean label plus a 0..1 score used in the composite
# priority score. Compound terms must be matched before their substrings.
# ---------------------------------------------------------------------------
ORDINAL_TABLE = [
    ("very high",    "Very High",   1.00),
    ("very low",     "Very Low",    0.10),
    ("medium-high",  "Medium-High", 0.75),
    ("medium-low",   "Medium-Low",  0.35),
    ("med-high",     "Medium-High", 0.75),
    ("moderate-high","Medium-High", 0.75),
    ("high",         "High",        0.85),
    ("medium",       "Medium",      0.50),
    ("moderate",     "Medium",      0.50),
    ("low",          "Low",         0.25),
]


def normalise_ordinal(raw):
    """Return (label, score) for a Low..Very High ordinal field.

    score is 0..1 (None when unscorable). Dashes are unified so en/em-dash
    variants collapse onto the hyphenated form before matching.
    """
    if raw is None:
        return None, None
    s = str(raw).strip().lower().replace("–", "-").replace("—", "-")
    s = re.sub(r"\s*-\s*", "-", s)        # "medium - high" -> "medium-high"
    s = re.sub(r"\s+", " ", s)
    if not s or s in ("none", "n/a", "na", "-", "unclear"):
        return None, None
    for kw, label, score in ORDINAL_TABLE:
        if kw in s:
            return label, score
    return None, None


# ---------------------------------------------------------------------------
# Intended Hub Role — free text, often blank, with typos ("Oerational").
# Collapse to a single primary category; hybrids resolve by priority
# (Hub-native > Federation > Derived > Operational > Reference) because the
# most operationally consequential label should drive sequencing.
# ---------------------------------------------------------------------------
def normalise_role(raw):
    if raw is None:
        return "Unspecified"
    s = str(raw).strip().lower()
    if not s or s in ("none", "-", "`", "n/a", "na"):
        return "Unspecified"
    if "hub-native" in s or "hub native" in s or "hubnative" in s:
        return "Hub-native"
    if "federation" in s or "federate" in s:
        return "Federation"
    if "derived" in s:
        return "Derived"
    if "oerational" in s or "operational" in s:
        return "Operational"
    if "reference" in s:
        return "Reference"
    # Long free-text with no role keyword = misplaced description, not a role.
    return "Unspecified"


# ---------------------------------------------------------------------------
# Access status + file format -> integration signals for the data team.
# ---------------------------------------------------------------------------
def normalise_access(raw):
    if raw is None:
        return "Unknown"
    s = str(raw).strip().lower()
    if not s or s in ("none", "-", "n/a", "na", "unknown"):
        return "Unknown"
    if any(k in s for k in ("open", "public", "cc-by", "cc by", "free")):
        return "Open"
    if any(k in s for k in ("restrict", "embargo", "request", "closed",
                            "private", "internal", "license", "licence",
                            "permission", "negoti")):
        return "Restricted"
    return "Unknown"


CLOUD_FORMAT_KW = ("cog", "cloud-optim", "cloud optim", "zarr", "parquet",
                   "geoparquet", "stac", "arrow", "api", "rest", "ogc",
                   "wms", "wfs", "wcs")
OPEN_STD_KW = ("geotiff", "tif", "netcdf", ".nc", "csv", "json", "geojson",
               "shapefile", "shp", "xlsx", "excel", "hdf", "sqlite", "geopackage")


def classify_format(raw):
    """Return (format_class, api_or_cloud:bool). Heuristic from file format text."""
    if raw is None:
        return "Unknown", False
    s = str(raw).strip().lower()
    if not s or s in ("none", "-", "n/a", "na"):
        return "Unknown", False
    if any(k in s for k in CLOUD_FORMAT_KW):
        return "Cloud-optimised", True
    if any(k in s for k in OPEN_STD_KW):
        return "Standard-open", False
    return "Proprietary/Other", False


def integration_hint(access_norm, format_class, api_or_cloud):
    """Heuristic federation-vs-ingest signal for the data team. Advisory only."""
    if access_norm == "Restricted":
        return "Negotiate access"
    if api_or_cloud:
        return "Federate — ready"
    if access_norm == "Open" and format_class == "Standard-open":
        return "Federate or light ingest"
    if access_norm == "Open":
        return "Ingest candidate"
    return "Assess"


# ---------------------------------------------------------------------------
# Climate-input extraction — supports the strategy's stated purpose of
# detecting duplication/dependency across pipelines. Free-text Primary Climate
# Inputs are scanned for recognised dataset names. Order: longer/specific first.
# ---------------------------------------------------------------------------
CLIMATE_INPUT_PATTERNS = [
    (r"\bag-?era5\b",                 "AgERA5"),
    (r"\bera5\b",                     "ERA5"),
    (r"\bera-?interim\b",             "ERA-Interim"),
    (r"\bchirps\b",                   "CHIRPS"),
    (r"\bchirts\b",                   "CHIRTS"),
    (r"\bcmip-?6\b",                  "CMIP6"),
    (r"\bcmip-?5\b",                  "CMIP5"),
    (r"\bcordex\b",                   "CORDEX"),
    (r"\bnex-?gddp\b",                "NEX-GDDP"),
    (r"\bagmerra\b",                  "AgMERRA"),
    (r"\bmerra-?2?\b",                "MERRA-2"),
    (r"\bterraclimate\b",             "TerraClimate"),
    (r"\bworldclim\b",                "WorldClim"),
    (r"\bchelsa\b",                   "CHELSA"),
    (r"\bcru\b",                      "CRU TS"),
    (r"\bnasa\s*power\b",             "NASA POWER"),
    (r"\bgpm\b|\bimerg\b",            "GPM/IMERG"),
    (r"\btrmm\b",                     "TRMM"),
    (r"\bmodis\b",                    "MODIS"),
    (r"\bsentinel\b",                 "Sentinel"),
    (r"\blandsat\b",                  "Landsat"),
    (r"\bisimip\b",                   "ISIMIP"),
    (r"\benacts\b",                   "ENACTS"),
    (r"\bagcfsr\b",                   "AgCFSR"),
    (r"\bstation\b|\bgauge\b|\bobserv",   "Station / observational"),
    (r"\breanalys",                   "Reanalysis (unspecified)"),
]


def extract_climate_inputs(raw):
    """Return a sorted list of recognised climate input datasets from free text."""
    if not raw:
        return []
    low = str(raw).lower()
    found = []
    for pat, label in CLIMATE_INPUT_PATTERNS:
        if re.search(pat, low) and label not in found:
            found.append(label)
    return found


EMAIL_RE = re.compile(r"[\w.+-]+@[\w-]+\.[\w.-]+")


def extract_email(raw):
    """Pull an email out of a 'Name (email@x)' contact string.

    De-obfuscates common anti-scrape forms ('foo (at) bar (dot) org').
    """
    if raw is None:
        return None
    s = str(raw)
    s = re.sub(r"\s*[\(\[\{]\s*at\s*[\)\]\}]\s*", "@", s, flags=re.I)
    s = re.sub(r"\s*[\(\[\{]\s*dot\s*[\)\]\}]\s*", ".", s, flags=re.I)
    m = EMAIL_RE.search(s)
    return m.group(0).rstrip(".").lower() if m else None


def yes_no(raw):
    if raw is None:
        return None
    s = str(raw).strip().lower()
    if s.startswith(("y", "true", "1")):
        return True
    if s.startswith(("n", "false", "0")):
        return False
    return None


# ---------------------------------------------------------------------------
# Composite priority score — single shared definition.
# Each component is 0..1; missing components drop out of the weighted average
# so absent data lowers confidence, not the score. rank_score is filled in a
# second pass (it is centre-relative). Mirror these weights in the dashboard.
# ---------------------------------------------------------------------------
DEFAULT_WEIGHTS = {
    "decision_relevance": 0.20,
    "reuse_potential":    0.15,
    "technical_readiness":0.20,
    "contemporary_validity":0.10,
    "sustainability":     0.10,
    "rank":               0.15,
    "role":               0.10,
}

ROLE_SCORE = {
    "Hub-native": 1.00,
    "Derived":    0.80,
    "Federation": 0.70,
    "Reference":  0.60,
    "Operational":0.60,
    "Unspecified":None,
}


def composite_score(components):
    """Weighted average of available 0..1 components -> 0..100, or None."""
    num = den = 0.0
    for key, w in DEFAULT_WEIGHTS.items():
        v = components.get(key)
        if v is not None:
            num += w * v
            den += w
    return round(100 * num / den) if den else None


# ---------------------------------------------------------------------------
# Sheet readers
# ---------------------------------------------------------------------------
def find_header_row(ws, min_cols=3):
    """Return the 0-based row index of the first row with >= min_cols non-None values."""
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if sum(1 for c in row if c is not None) >= min_cols:
            return i
    return 0


SKIP_NAMES = {"asset name", "unknown", ""}

def sheet_to_rows(ws) -> tuple[list[str], list[dict]]:
    """Return (headers, data_rows) from a sheet, skipping header and EXAMPLE rows.
    data_rows is a list of dicts; order matches asset order in the sheet."""
    header_row_idx = find_header_row(ws)
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(c).strip() if c else f"col{i}" for i, c in enumerate(rows[header_row_idx])]
    data = []
    for row in rows[header_row_idx + 1:]:
        key = str(row[0]).strip() if row[0] else ""
        if key.lower() in SKIP_NAMES:
            continue
        if key.upper().startswith("EXAMPLE"):
            continue
        data.append(dict(zip(headers, row)))
    return headers, data


def collect_example_values() -> set:
    """Scan every submission for template EXAMPLE-row cell text.

    The template ships two worked examples (Yield Potential, MapSPAM). Some
    submitters delete the example rows but leave the copied boilerplate in their
    real assets, so the blocklist must be built across ALL workbooks, not just
    the one being read.
    """
    vals = set()
    for filename in CENTRE_FILES.values():
        path = SUBMISSIONS_DIR / filename
        if not path.exists():
            continue
        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        except Exception:
            continue
        for sn in wb.sheetnames:
            for row in wb[sn].iter_rows(values_only=True):
                if row and row[0] and str(row[0]).strip().upper().startswith("EXAMPLE"):
                    for cell in row[1:]:
                        if cell and len(str(cell).strip()) > 20:
                            vals.add(" ".join(str(cell).split()).lower())
        wb.close()
    return vals


def read_centre(centre: str, filepath: Path, example_vals: set = None) -> list[dict]:
    """Read one centre's Excel file and return a list of asset dicts."""
    example_vals = example_vals or set()
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"  WARNING: Could not open {filepath.name}: {e}", file=sys.stderr)
        return []

    sheets = {s.lower(): s for s in wb.sheetnames}

    def get_ws(name):
        actual = sheets.get(name.lower())
        return wb[actual] if actual else None

    ws_a  = get_ws(SHEET_A)
    ws_b  = get_ws(SHEET_B)
    ws_c  = get_ws(SHEET_C)
    ws_d  = get_ws(SHEET_D)
    ws_e  = get_ws(SHEET_E)
    ws_fg = get_ws(SHEET_FG)

    if not ws_a:
        print(f"  WARNING: No '{SHEET_A}' sheet in {filepath.name}", file=sys.stderr)
        return []

    _, a_rows  = sheet_to_rows(ws_a)
    _, b_rows  = sheet_to_rows(ws_b)  if ws_b  else ([], [])
    _, c_rows  = sheet_to_rows(ws_c)  if ws_c  else ([], [])
    _, d_rows  = sheet_to_rows(ws_d)  if ws_d  else ([], [])
    _, e_rows  = sheet_to_rows(ws_e)  if ws_e  else ([], [])
    _, fg_rows = sheet_to_rows(ws_fg) if ws_fg else ([], [])

    def is_boilerplate(v):
        return v is not None and " ".join(str(v).split()).lower() in example_vals

    # Join sheets by row index (all sheets share the same asset order)
    def at(rows, i):
        return rows[i] if i < len(rows) else {}

    assets = []
    for i, a in enumerate(a_rows):
        b  = at(b_rows,  i)
        c  = at(c_rows,  i)
        d  = at(d_rows,  i)
        e  = at(e_rows,  i)
        fg = at(fg_rows, i)

        # Pull raw values — search all sheets, matching column header case-insensitively
        def get(*keys):
            for k in keys:
                kl = k.lower().strip()
                for sheet_row in [a, b, c, d, e, fg]:
                    for col, val in sheet_row.items():
                        if col.lower().strip() == kl and val is not None:
                            return val
            return None

        name        = str(a.get("Asset Name", "") or list(a.values())[0] or "").strip()
        raw_domain  = get("Climate Domain", "climate_domain", "Domain")
        raw_type    = get("Asset Type", "asset_type", "Asset_Type")
        raw_geo     = get("Spatial Coverage", "spatial_coverage")
        raw_rank    = get("Asset Rank", "asset_rank")
        raw_role    = get("Intended Hub Role", "hub_role", "Hub Role")
        raw_access  = get("Access Status", "access_status")
        raw_contact = get("Primary Contact", "primary_contact")
        raw_inputs  = get("Primary Climate Inputs", "primary_climate_inputs")
        nr_lbl, _   = normalise_ordinal(get("National_Relevance", "National Relevance"))
        raw_format  = get("Primary File Format", "File Format", "primary_file_format")
        raw_maint   = get("Actively Maintained", "actively_maintained")
        raw_found   = get("Foundational_to_Work", "Foundational to Work")

        def clean(v):
            s = str(v).strip() if v is not None else None
            return None if s in (None, "None", "-", "N/A", "NA", "") else s

        def clean_text(v):
            """clean() plus boilerplate scrub for free-text fields."""
            s = clean(v)
            return None if is_boilerplate(s) else s

        # Ordinal quality scores -> clean label + 0..1 score
        dr_lbl,  dr_sc  = normalise_ordinal(get("Decision Relevance", "decision_relevance"))
        ru_lbl,  ru_sc  = normalise_ordinal(get("Reuse Potential", "reuse_potential"))
        tr_lbl,  tr_sc  = normalise_ordinal(get("Technical Readiness", "technical_readiness"))
        cv_lbl,  cv_sc  = normalise_ordinal(get("Contemporary Validity", "contemporary_validity"))
        su_lbl,  su_sc  = normalise_ordinal(get("Sustainability"))

        role_norm   = normalise_role(raw_role)
        access_norm = normalise_access(raw_access)
        fmt_class, api_or_cloud = classify_format(raw_format)

        try:
            rank_num = int(float(str(raw_rank).strip())) if clean(raw_rank) else None
        except (ValueError, TypeError):
            rank_num = None

        components = {
            "decision_relevance":   dr_sc,
            "reuse_potential":      ru_sc,
            "technical_readiness":  tr_sc,
            "contemporary_validity":cv_sc,
            "sustainability":       su_sc,
            "role":                 ROLE_SCORE.get(role_norm),
            "rank":                 None,   # filled in centre-relative pass
        }

        assets.append({
            "centre":               centre,
            "hub_funded":           centre in HUB_FUNDED,
            "name":                 name,
            "nominator":            clean(get("Nominating_Person", "Nominator")),
            "asset_type":           clean(raw_type),
            # short_description is NOT boilerplate-scrubbed: the only match is
            # MapSPAM, which IS the template example but also a real IFPRI asset,
            # so keeping its description is correct and harms no other asset.
            "short_description":    clean(get("Short Description", "Short_Description")),
            "climate_domain":       clean(raw_domain),
            "farming_system":       clean(get("Farming System", "farming_system")),
            "commodity":            clean(get("Commodity")),
            "spatial_coverage":     clean(raw_geo),
            "spatial_resolution":   clean(get("Spatial Resolution", "spatial_resolution")),
            "temporal_type":        clean(get("Temporal Type", "temporal_type")),
            "year_last_updated":    clean(get("Year Last Updated", "year_last_updated")),
            "decision_relevance":   clean(get("Decision Relevance", "decision_relevance")),
            "reuse_potential":      clean(get("Reuse Potential", "reuse_potential")),
            "technical_readiness":  clean(get("Technical Readiness", "technical_readiness")),
            "contemporary_validity":clean(get("Contemporary Validity", "contemporary_validity")),
            "sustainability":       clean(get("Sustainability")),
            "asset_rank":           clean(raw_rank),
            "hub_role":             clean(raw_role),
            # Newly ingested context fields
            "access_status":        clean(raw_access),
            "url":                  clean(get("URL or Repository", "URL", "Repository")),
            "file_format":          clean(raw_format),
            "data_structure_type":  clean(get("Data Structure Type", "data_structure_type")),
            "primary_use_case":     clean_text(get("Primary Use Case", "primary_use_case")),
            "justification":        clean_text(get("Justification")),
            "primary_contact":      clean(raw_contact),
            "contact_email":        extract_email(raw_contact) or extract_email(get("Nominating_Person", "Nominator")),
            "asset_organization":   clean(get("Asset_Organization", "Asset Organization")),
            # Strategy-aligned context fields
            "primary_climate_inputs": clean(raw_inputs),
            "climate_inputs_norm":  extract_climate_inputs(raw_inputs),
            "output_variable_type": clean(get("Output Variable Type", "output_variable_type")),
            "user_groups":          clean(get("User Groups", "user_groups")),
            "cgiar_programs":       clean(get("CGIAR Programs Directly Using Asset")),
            "partners":             clean(get("Partners or Projects Using Asset")),
            "national_relevance":   nr_lbl,
            "actively_maintained":  yes_no(raw_maint),
            "foundational":         yes_no(raw_found),
            # Normalised category fields
            "domain_norm":          normalise_domain(str(raw_domain) if raw_domain else ""),
            "type_norm":            normalise_type(raw_type),
            "geo_norm":             normalise_geo(str(raw_geo) if raw_geo else ""),
            # Normalised ordinal labels + 0..1 scores
            "decision_relevance_norm":  dr_lbl,
            "reuse_potential_norm":     ru_lbl,
            "technical_readiness_norm": tr_lbl,
            "contemporary_validity_norm":cv_lbl,
            "sustainability_norm":      su_lbl,
            # Integration signals
            "hub_role_norm":        role_norm,
            "access_norm":          access_norm,
            "format_class":         fmt_class,
            "api_or_cloud":         api_or_cloud,
            "integration_hint":     integration_hint(access_norm, fmt_class, api_or_cloud),
            # Priority scoring
            "asset_rank_num":       rank_num,
            "score_components":     components,
        })

    return assets


def apply_merge_log(assets: list[dict], merge_log: dict) -> list[dict]:
    """Apply consolidations from merge_log['applied'] to asset list.

    For each consolidation: rename the first matching original to consolidated_name,
    then remove the remaining originals.
    """
    to_remove = set()
    for merge in merge_log.get("applied", []):
        centre       = merge["centre"]
        originals    = merge["merged_names"]      # ordered list
        consolidated = merge["consolidated_name"]
        found = [a for a in assets if a["centre"] == centre and a["name"] in originals]
        if not found:
            print(f"  WARNING: no assets matched merge '{consolidated}' for {centre}",
                  file=sys.stderr)
            continue
        # Rename the first found asset; mark the rest for removal
        found[0]["name"] = consolidated
        for a in found[1:]:
            to_remove.add((a["centre"], a["name"]))
        print(f"  Merged {len(found)} → 1: {consolidated} ({centre})")

    return [a for a in assets if (a["centre"], a["name"]) not in to_remove]


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    print("Reading submissions...")
    example_vals = collect_example_values()
    print(f"  Template boilerplate signatures: {len(example_vals)}")
    all_assets = []
    for centre, filename in CENTRE_FILES.items():
        path = SUBMISSIONS_DIR / filename
        if not path.exists():
            print(f"  MISSING: {path.name} — skipping {centre}")
            continue
        assets = read_centre(centre, path, example_vals)
        print(f"  {centre}: {len(assets)} assets")
        all_assets.extend(assets)

    print(f"\nRaw total: {len(all_assets)}")

    # Centre-level geo fallbacks: where spatial_coverage is blank but centre mandate is clear
    CENTRE_GEO_FALLBACK = {
        "AfricaRice": "Africa",   # Africa Rice Center — all field work is Africa
    }
    for a in all_assets:
        if a["geo_norm"] == "Not specified" and a["centre"] in CENTRE_GEO_FALLBACK:
            a["geo_norm"] = CENTRE_GEO_FALLBACK[a["centre"]]

    # Apply merge log
    with open(MERGE_LOG_PATH) as f:
        merge_log = json.load(f)
    all_assets = apply_merge_log(all_assets, merge_log)
    print(f"After consolidation: {len(all_assets)}")

    # ----------------------------------------------------------------------
    # Centre-relative rank: Asset Rank is ranked 1..N *within each centre*,
    # not across the system. Convert to a 0..1 score (rank 1 = best) and a
    # top-3 flag, then compute the shared composite priority score.
    # ----------------------------------------------------------------------
    centre_max_rank = defaultdict(int)
    for a in all_assets:
        if a["asset_rank_num"]:
            centre_max_rank[a["centre"]] = max(centre_max_rank[a["centre"]],
                                               a["asset_rank_num"])
    for a in all_assets:
        r, mx = a["asset_rank_num"], centre_max_rank[a["centre"]]
        if r and mx:
            a["rank_score"] = round((mx - r + 1) / mx, 3)
            a["is_top3_in_centre"] = r <= 3
        else:
            a["rank_score"] = None
            a["is_top3_in_centre"] = False
        a["score_components"]["rank"] = a["rank_score"]
        a["priority_score"] = composite_score(a["score_components"])

    n_scored = sum(1 for a in all_assets if a["priority_score"] is not None)
    print(f"Priority score computed for {n_scored}/{len(all_assets)} assets")

    # ----------------------------------------------------------------------
    # Apply researched contact-email overrides for nominators who left the
    # field blank. Auditable file, never hand-edit assets.json. Each override
    # carries a confidence + source so the dashboard can flag inferred ones.
    # ----------------------------------------------------------------------
    overrides_path = ROOT / "data" / "contact_overrides.json"
    if overrides_path.exists():
        def norm_name(v):
            return " ".join(str(v).split()).lower() if v else ""

        with open(overrides_path) as f:
            overrides = json.load(f).get("overrides", [])
        idx = {}
        for o in overrides:
            for name in o.get("match", []):
                idx[norm_name(name)] = o
        applied = 0
        for a in all_assets:
            if a.get("contact_email"):
                continue
            # Prefer the named primary contact, then the nominator.
            o = idx.get(norm_name(a.get("primary_contact"))) or idx.get(norm_name(a.get("nominator")))
            if o and o.get("email"):
                a["contact_email"] = o["email"]
                a["contact_email_confidence"] = o.get("confidence", "inferred")
                a["contact_email_source"] = o.get("source")
                applied += 1
        print(f"Applied {applied} contact-email overrides")

    # Write output
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_PATH, "w") as f:
        json.dump(all_assets, f, indent=2, default=str)
    print(f"\nWritten: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
